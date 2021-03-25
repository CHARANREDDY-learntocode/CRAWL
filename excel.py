import threading
import os
import zipfile

import paho.mqtt.client as paho
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle,colors,Alignment,Font,PatternFill,Border,Side
from openpyxl.chart import BarChart, Reference

def previous_data(data):
    #get the data from the mqtt file response
    print(f'before: \n{data}')
    data_list = [node.split('/')[:2] for node in data.strip().split(' ')]
    write_data(data_list)

def current_data(data):
    #get the data from the simulate file response
    print(f'After: \n{data}')
    data_list = [node.split('/')[:2] for node in data.strip().split(' ')]
    write_data(data_list, before= False)

def write_data(data, before=True):
    #write the data to excel file
    try: wb = load_workbook(filename='report.xlsx')
    except FileNotFoundError: wb = Workbook()
    except:
        try:
            os.remove('report.xlsx')
            wb = Workbook()
        except: wb = Workbook()
    ws = wb.active
    ws.title = 'Nodes Data Table'
    counter = 0
    headings = ['Sl.No', 'Node', 'Before', 'After']
    for column in range(1, 5): ws.cell(row= 1, column= column, value = headings[column-1])
    if before:
        for col in ws.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=3):
            iter = 0
            for cell in col:
                if iter == 0: cell.value = counter + 1
                elif iter == 1: cell.value = data[counter][0]
                elif iter == 2: cell.value = float(data[counter][1])
                iter += 1
            counter += 1
        #create bar chart for the before data
        try: ws1 = wb['Before & After']
        except: 
            ws1= wb.create_sheet('Before & After')
            ws1.title = 'Before & After'
        deficit, sufficient, surplus = 0, 0, 0
        for node in data:
            if float(node[1]) < 40: deficit += 1
            elif float(node[1]) < 70: sufficient += 1
            else: surplus += 1
        head = ['Deficit', 'Sufficient', 'Surplus']
        lst = [deficit, sufficient, surplus]
        cell = ws1.cell(row=1, column=1)
        cell.value = 'Before'
        for col in range(2, 5): ws1.cell(row= 1, column= col, value = head[col-2])
        for col in range(2, 5): ws1.cell(row= 2, column= col, value = lst[col-2])
        ws1.merge_cells(start_row=1,end_row=2,end_column=1,start_column=1)
        # create data for plotting
        values = Reference(ws1, min_col = 2, min_row = 1,
                                max_col = 4, max_row = 2)
        # Create object of BarChart class
        chart = BarChart()
        chart.x_axis.delete = False
        # adding data to the Bar chart object
        chart.add_data(values, titles_from_data= True)

        # set the title of the chart
        chart.title = " The Nodes Count Before Pipelining "

        # set the title of the x-axis
        chart.x_axis.title = "Node Type"

        # set the title of the y-axis
        chart.y_axis.title = "Number of Nodes"
        ws1.add_chart(chart, "B10")
    else:
        for col in ws.iter_rows(min_row=2, max_row=len(data)+1, min_col=4, max_col=4): 
                for cell in col: cell.value =  float(data[counter][1])
                counter += 1
        #create bar chart for the before data
        try: ws1 = wb['Before & After']
        except: 
            ws1= wb.create_sheet('Before & After')
            ws1.title = 'Before & After'
        deficit, sufficient, surplus = 0, 0, 0
        for node in data:
            if float(node[1]) < 40: deficit += 1
            elif float(node[1]) < 70: sufficient += 1
            else: surplus += 1
        head = ['deficit', 'sufficient', 'surplus']
        lst = [deficit, sufficient, surplus]
        cell = ws1.cell(row=1, column=5)
        cell.value = 'After'
        for col in range(6, 9): ws1.cell(row= 1, column= col, value = head[col-6])
        for col in range(6, 9): ws1.cell(row= 2, column= col, value = lst[col-6])
        # create data for plotting
        ws1.merge_cells(start_row=1,end_row=2,end_column=5,start_column=5)
        values = Reference(ws1, min_col = 6, min_row = 1,
                                max_col = 8, max_row = 2)
        # Create object of BarChart class
        chart = BarChart()
        chart.x_axis.delete = False
        # adding data to the Bar chart object
        chart.add_data(values, titles_from_data=True)

        # set the title of the chart
        chart.title = " The Nodes Count After Pipelining "

        # set the title of the x-axis
        chart.x_axis.title = "Node Type"

        # set the title of the y-axis
        chart.y_axis.title = "Number of Nodes"
        # add chart to the sheet
        # the top-left corner of a chart
        # is anchored to cell E2 .
        ws1.add_chart(chart, "M10")
    try: wb.save('report.xlsx')
    except PermissionError: print('Please close the excel file, to continue logging the data.')  

def on_message(client, userdata, msg):
    #print the data when message arrives
    data = msg.payload.decode('utf-8')
    if msg.topic == 'NODE RESPONSE': threading.Thread(target=previous_data, args=[data]).start()
    else: threading.Thread(target= current_data, args=[data]).start()

def on_subscribe(client, userdata, mid, granted_qos):
    #print the success message when subscribed to a node
    print('subscribed')

def connect():
    #connect to the mqtt server and subscribe to the nodes
    global client
    client = paho.Client()
    client.on_message = on_message
    client.on_subscribe = on_subscribe
    client.username_pw_set("318624ea71ff415090c4c24ee85dc568")
    client.connect('broker.hivemq.com')
    client.subscribe('NODE RESPONSE', qos=1)
    client.subscribe('STABLE NODE RESPONSE', qos=1)
    client.loop_forever()

connect()